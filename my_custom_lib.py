"""This module converts an excel table to a markdown table. Numbers show a
    thousands separator and will be right aligned. A line separator before
    the total can be inserted.

The excel table is loaded into a dataframe. The thousands separator is
set to dot. The last line of the dataframe is duplicated. In the line before
last, a separator-line is inserted.

Args:
    file_name (str): the filename of the excel data file
    sheet_name (str): the sheetname in the excel file to load the data from
    table_name (str): the tablename in the excel sheet to load the data from

Example:
    xls2md("sia-projecten.xlsx", "ThemaSelect", "Thema")

Attributes:
    df is the Pandas DataFrame that the excel table is loaded into

To do:
     * generalize the module such that it accepts tables with multiple
        numeric columns maybe later pass a list of dictionary's where
        the dictionary holds {column name : sum=true} or {column name
        : sum = false}
     * include a test check_table_exists -> True/False
     * possibly use argparse as shown in excel2markdown by SpeerSec

.. _Ref excel2markdown by SpeerSec:
   https://github.com/SpeerSec/excel2markdown
"""

import sys
import locale
import pandas as pd
from pandas.api.types import is_numeric_dtype
from openpyxl import load_workbook


def xls2md(table: dict) -> pd.DataFrame:
    """
    Args:
        table
    """
    locale.setlocale(locale.LC_ALL, "de_DE")
    print(table)

    if table_exists(table["name"]):
        coordinaten = unpack_xy(table["range"])
        sia_projecten = pd.ExcelFile(table["file"])
        with sia_projecten as xls:
            df = pd.read_excel(
                xls,
                table["sheet"],
                skiprows=2,
                nrows=coordinaten[3] - coordinaten[2],
                usecols=coordinaten[0] + ":" + coordinaten[1],
            )
        # print(df[df.columns[1]])
        if is_numeric_dtype(df[df.columns[1]]):
            df[df.columns[1]] = df[df.columns[1]].apply(
                lambda x: locale.format_string(
                    "%10.0f", x, grouping=True, monetary=True
                )
            )
            df = df.fillna("")
            df = df.astype(str)
            df = pd.concat([df, df.iloc[-1:]], ignore_index=True)
            df.at[df.index[-2], df.columns[0]] = ""
            df.at[df.index[-2], df.columns[1]] = "-" * 7

        return df


# def tblname2df(filename: str, tablename: str) -> pd.DataFrame:
# wb = load_workbook(filename, data_only=True)
# ws = wb[sheetname]
# range of table
# cellrange = ws.tables[tablename].ref
# column range of table
# cols = [column.value for column in ws[cellrange][0]]
# number of rows in table
# n_rows = len(ws[cellrange][1:])
# number of rows to skip
# skip = int(cellrange[1]) - 1
# return the dataframe
# return pd.read_excel(filename, usecols=cols, skiprows=skip, nrows=n_rows)
# return pd.read_excel(filename, sheetname, usecols=cols, skiprows=skip,
# nrows=n_rows)


def sheet_exists(filename: str, sheet_name: str) -> bool:
    """
    open an Excel file and return a workbook
    file, sheet, table mogelijk in 1 functie samennemen? *kwars
    """
    wb = load_workbook(filename, read_only=True)
    if sheet_name not in wb.sheetnames:
        sys.exit("Error: sheet is not in Excel file")
        # The module shall have only one (normal) exit
        # I expect there shall be an error handler (function)
    return True


def table_exists(table: dict) -> bool:
    """
    Check file and sheet
    if file_exists(table['file']) and sheet_exists(table['sheet']):
    Check table
    not correct!
    """
    if table not in [1, 2]:
        sys.exit("Error: table not in Excel file")
    return True


def add_thousands_separator(df: pd.DataFrame, column: list[str]) -> pd.DataFrame:
    """
    To be added to all numeric columns, not just column[1]
    This will throw an error if there is a non-numeric in column Aantal
    """
    df[column[1]] = df[column[1]].apply(
        lambda x: locale.format_string("%10.0f", x, grouping=True, monetary=True)
    )
    return df


def add_separator_line(df: pd.DataFrame, column: list[str]) -> pd.DataFrame:
    """
    add a separator line above the totals
    """
    df = df.astype({column[0]: str, column[1]: str})
    df = pd.concat([df, df.iloc[-1:]], ignore_index=True)
    df.at[df.index[-2], column[0]] = ""
    df.at[df.index[-2], column[1]] = "-" * 7
    return df


def unpack_xy(cell_range: str) -> list:
    """
    to be improved as currently constrained to A:Z and 1:99
    """
    x1, x2, y1, y2 = "", "", 0, 0
    x1 = cell_range[0:1]  # only works up to Z
    x2 = cell_range[3:4]  # only works up to Z
    y1 = int(cell_range[1:2])  # only works up to 99
    if len(cell_range) == 5:  # assuming:
        y2 = int(cell_range[4:5])  # row 1 and last row single digit
    elif len(cell_range) == 6:  # assuming:
        y2 = int(cell_range[4:6])  # row 1 single digit and last row double digit
    elif len(cell_range) == 7:  # assuming:
        y2 = int(cell_range[5:7])  # row 1 and last row double digit
    # print(x1, x2, y1, y2)
    return [x1, x2, y1, y2]
